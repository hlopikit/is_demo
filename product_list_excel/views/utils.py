from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def date_parse(products):

    dates_dict = dict()
    for product in products:
        dates_dict[product["ID"]] = product["DATE_CREATE"]
    for key, date in dates_dict.items():
        res = dt.strptime(date, "%Y-%m-%dT%H:%M:%S%z")
        formatted_date = res.strftime("%d.%m.%Y %H:%M:%S")
        dates_dict[key] = formatted_date

    return dates_dict


def users_in_dict(users):

    users_dict = dict()
    for user in users:
        users_dict[user["ID"]] = f'{user["NAME"]} {user["LAST_NAME"]}'

    return users_dict


def save_file(products, users):

    # Создание нового файла Excel
    workbook = Workbook()
    sheet = workbook.active
    users_name = users_in_dict(users)
    date = date_parse(products)
    # Запись данных в ячейки
    data = [
        ["ID", "Название", "Код товара", "Дата создания", "Кем изменено", "Кем создано", "ID каталога",
         "Описание товара", "Цена", "Валюта"],
    ]

    for product in products:
        data.append([product['ID'], product['NAME'], product['CODE'], date[product['ID']],
                     users_name[product['MODIFIED_BY']], users_name[product['CREATED_BY']], product['CATALOG_ID'],
                     product['DESCRIPTION'], product['PRICE'], product['CURRENCY_ID']])

    for row in data:
        sheet.append(row)

    # Определение диапазона данных
    num_rows = len(data)
    num_cols = len(data[0])
    start_cell = "A1"
    end_cell = get_column_letter(num_cols) + str(num_rows + 1)

    # Создание объекта таблицы
    table = Table(displayName="Table1", ref=f"{start_cell}:{end_cell}")

    # Определение стиля таблицы
    table_style = TableStyleInfo(name="TableStyleMedium6", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True,
                                 showColumnStripes=False)

    table.tableStyleInfo = table_style

    # Добавление таблицы на лист
    sheet.add_table(table)

    # Автоматическое расширение столбцов
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Оптимизированное автоматическое расширение столбцов
    sheet.calculate_dimension()

    # Сохранение файла
    workbook.save('example.xlsx')
