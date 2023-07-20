from openpyxl import load_workbook

OBJECT_CRM = {"Лиды": 1,
              "Сделки": 2,
              "Контакты": 3,
              "Компании": 4,
              "Коммерческие предложения": 7,
              "Новые счета": 31}


def excel_to_dict(file_path, sheet_name):
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]

    data = list()
    address_company = list()
    list_company = list()
    origin_dict = dict()

    headers = [cell.value for cell in sheet[1]]
    origin_id = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if sheet_name == "Контакты":
            row_data["PHONE"] = [{"VALUE": str(row_data["PHONE"]), "VALUE_TYPE": "WORK"}]
        elif sheet_name == "Компании":
            address_company.append({"ADDRESS_1": str(row_data["ADDRESS"]), "CITY": str(row_data["ADDRESS_CITY"])})
            origin_dict[origin_id] = None
        data.append(row_data)
        origin_id += 1

    list_company.append(data)
    list_company.append(origin_dict)
    list_company.append(address_company)

    return list_company


def get_sheet_names(file_path):
    workbook = load_workbook(filename=file_path)
    sheet_names = workbook.sheetnames

    return sheet_names
