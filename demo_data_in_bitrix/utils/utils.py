import time

from openpyxl import load_workbook
import pandas as pd

OBJECT_CRM = {"Лиды": 1,
              "Сделки": 2,
              "Контакты": 3,
              "Компании": 4,
              "Коммерческие предложения": 7,
              "Новые счета": 31}


def add_origin_prefix(data, prefix):
    """
    В файле ексель(гугл) док мы делаем связку менжду страницами по полю ORIGIN_ID, т.к при импорте в Б24 мы получим новые ID сущностей.
    Эта функая помогает делать уникальные ORIGIN_ID для единовремнного импорта демоданных
    для всех записей добавляет префикс (Мы его возьмем как текущее время с микросекундами)
    у все хаписей "10" стенет "1690205018.084936_10"
    :param data:
    :param prefix:
    :return:
    """
    for d in data:
        # Добавляем префикс для ORIGIN_ID
        if d.get('ORIGIN_ID'):
            d['ORIGIN_ID'] = "{}_{}".format(prefix, d.get('ORIGIN_ID'))
    return data


def make_links_from_origin(data, excel_field, b24_field, source_map, prefix):
    # Принимает список из листов экселя
    # заменяет excel_field=COMPANY_ORIGIN_ID на b24_field=COMPANY_ID учитывая предыдущую замену при импорте демоданных
    # для создания правльной адресации на только что созданные сущности
    for d in data:
        # Добавляем префикс для ORIGIN_ID
        if d.get(excel_field):
            d[b24_field] = \
                source_map["{}_{}".format(prefix, d.get(excel_field))]['ID']
    return data


def load_crm(crm_items, but, type_id):
    # выгружает элменты crm в битрикс
    methods = []
    for item in crm_items:
        methods.append(('crm.item.batchImport',
                        {"entityTypeId": type_id, "data": [item]}))
    but.batch_api_call(methods)


def import_data_from_xls(filename, but):
    # Принимает токен и excel файл с несколькоми страницами и загружает их в Битрикс24
    # sheet_names = get_sheet_names(filename)
    excel_file = pd.ExcelFile(filename)
    # для всего пакета загрузки будет одинаковый префикс у ORIGIN_ID
    load_origin_id_prefix = time.time()
    object_count = {"Лиды": None,
                    "Сделки": None,
                    "Контакты": None,
                    "Компании": None}

    if "Загружаем компании":  # Всегда True
        company_data = excel_file.parse('Компании').to_dict("records")
        company_data = add_origin_prefix(company_data, load_origin_id_prefix)
        object_count["Компании"] = len(company_data)
        load_crm(company_data, but, "4")

        companies = but.call_list_method('crm.company.list', {
            "SELECT": ["ORIGIN_ID", "ID"],
            "FILTER": {"%ORIGIN_ID": "{}_".format(load_origin_id_prefix)}})
        # Используем прием конвертации в адресный dict https://it-solution.kdb24.ru/article/218199/
        # после этого легко найдем companies_origin_id_dict['1690182208.5614886_7']['ID']
        companies_origin_id_dict = {item['ORIGIN_ID']: item for item in
                                    companies}
        for d in company_data:
            # https://dev.1c-bitrix.ru/rest_help/crm/requisite/methods/crm_address_add.php
            but.call_api_method("crm.address.add", {"fields": {
                "TYPE_ID": "1",  # Фактический адрес?
                "ENTITY_TYPE_ID": "4",  # 4 - для Компаний
                "ENTITY_ID": companies_origin_id_dict[d['ORIGIN_ID']]['ID'],
                "CITY": d["ADDRESS_CITY"],
                "ADDRESS_1": d["ADDRESS"],
            }})

    if "Загружаем контакты":
        contacts_data = excel_file.parse('Контакты').to_dict("records")
        contacts_data = add_origin_prefix(contacts_data, load_origin_id_prefix)
        contacts_data = make_links_from_origin(contacts_data,
                                               'COMPANY_ORIGIN_ID',
                                               'COMPANY_ID',
                                               companies_origin_id_dict,
                                               load_origin_id_prefix)
        for c in contacts_data:
            c["PHONE"] = [{"VALUE": str(c["PHONE"]), "VALUE_TYPE": "WORK"}]

        object_count["Контакты"] = len(contacts_data)
        load_crm(contacts_data, but, "3")

    if "Загружаем сделки":
        deals_data = excel_file.parse('Сделки').to_dict("records")
        object_count["Сделки"] = len(deals_data)
        load_crm(deals_data, but, "2")

    if "Загружаем лиды":
        leads_data = excel_file.parse('Лиды').to_dict("records")
        object_count["Лиды"] = len(leads_data)
        load_crm(leads_data, but, "1")

    return object_count


def excel_to_dict(file_path, sheet_name):
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]

    data = list()
    address_company = list()
    list_company = list()
    origin_dict = dict()

    headers = [cell.value for cell in sheet[1]]
    origin_id = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if sheet_name == "Контакты":
            row_data["PHONE"] = [
                {"VALUE": str(row_data["PHONE"]), "VALUE_TYPE": "WORK"}]
        elif sheet_name == "Компании":
            address_company.append({"ADDRESS_1": str(row_data["ADDRESS"]),
                                    "CITY": str(row_data["ADDRESS_CITY"])})
            origin_dict[origin_id] = None
        data.append(row_data)
        origin_id += 1

    list_company.append(data)
    list_company.append(origin_dict)
    list_company.append(address_company)

    return list_company


def get_sheet_names(file_path):
    workbook = load_workbook(filename=file_path)
    sheet_names = workbook.sheetnames

    return sheet_names
