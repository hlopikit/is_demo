from openpyxl import load_workbook
from copy import deepcopy as new

def add_origin_prefix(data, prefix):
    """
    В файле ексель(гугл) док мы делаем связку менжду страницами по полю ORIGIN_ID, т.к при импорте в Б24 мы получим новые ID сущностей.
    Эта функая помогает делать уникальные ORIGIN_ID для единовремнного импорта демоданных
    для всех записей добавляет префикс (Мы его возьмем как текущее время с микросекундами)
    у все хаписей "10" стенет "1690205018.084936_10"
    :param data:
    :param prefix:
    :return:
    """
    for d in data:
        # Добавляем префикс для ORIGIN_ID
        if d.get('ORIGIN_ID'):
            d['ORIGIN_ID'] = "{}_{}".format(prefix, d.get('ORIGIN_ID'))
    return data


def make_links_from_origin(data, excel_field, b24_field, source_map, prefix):
    # Принимает список из листов экселя
    # заменяет excel_field=COMPANY_ORIGIN_ID на b24_field=COMPANY_ID учитывая предыдущую замену при импорте демоданных
    # для создания правльной адресации на только что созданные сущности
    for d in data:
        # Добавляем префикс для ORIGIN_ID
        if d.get(excel_field):
            d[b24_field] = \
                source_map["{}_{}".format(prefix, d.get(excel_field))]['ID']
    return data


def load_crm(crm_items, but, type_id):
    # выгружает элементы crm в битрикс

    methods = []
    method = ('crm.item.batchImport', {'entityTypeId': type_id, 'data': []})
    counter = 0
    methods.append(new(method))
    for item in crm_items:
        counter += 1
        methods[-1][1]['data'].append(item)
        if counter % 20 == 0:
            methods.append(new(method))
    print(len(methods))
    but.batch_api_call(methods, timeout=300)


def excel_to_dict(file_path, sheet_name):
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]

    data = list()
    address_company = list()
    list_company = list()
    origin_dict = dict()

    headers = [cell.value for cell in sheet[1]]
    origin_id = 1
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if sheet_name == "Контакты":
            row_data["PHONE"] = [
                {"VALUE": str(row_data["PHONE"]), "VALUE_TYPE": "WORK"}]
        elif sheet_name == "Компании":
            address_company.append({"ADDRESS_1": str(row_data["ADDRESS"]),
                                    "CITY": str(row_data["ADDRESS_CITY"])})
            origin_dict[origin_id] = None
        data.append(row_data)
        origin_id += 1

    list_company.append(data)
    list_company.append(origin_dict)
    list_company.append(address_company)

    return list_company


def get_sheet_names(file_path):
    workbook = load_workbook(filename=file_path)
    sheet_names = workbook.sheetnames

    return sheet_names
